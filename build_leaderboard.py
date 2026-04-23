#!/usr/bin/env python3
"""
build_leaderboard.py — Render the scoring leaderboard xlsx as a styled HTML
page for the static site.

Reads the most recent leaderboard_YYYYMMDD.xlsx from scoring/output/ and
produces leaderboard.html with both the 365-day and 180-day rankings.

Usage:
    python3 build_leaderboard.py                    # writes site/leaderboard.html
    python3 build_leaderboard.py --out docs/        # custom output dir
"""

import argparse
import glob
import os
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

from scoring.benchmarks import BENCHMARK_TICKERS, all_benchmark_returns


PROJ = Path(__file__).resolve().parent
SCORING_OUTPUT = PROJ / "scoring" / "output"


def find_latest_leaderboard():
    """Find the most recent leaderboard xlsx."""
    pattern = str(SCORING_OUTPUT / "leaderboard_*.xlsx")
    files = sorted(glob.glob(pattern), reverse=True)
    if not files:
        return None
    return files[0]


def read_sheet(wb, sheet_name, max_rows=25):
    """Read a sheet into a list of dicts (header row as keys)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []
    headers = list(rows[0])
    data = []
    for row in rows[1:max_rows + 1]:
        data.append(dict(zip(headers, row)))
    return headers, data


def fmt_pct(val):
    if val is None:
        return "—"
    return f"{val * 100:.1f}%"


def fmt_float(val, decimals=2):
    if val is None:
        return "—"
    return f"{val:.{decimals}f}"


def party_badge(party):
    if party and "Democrat" in party:
        return '<span class="badge dem">D</span>'
    elif party and "Republican" in party:
        return '<span class="badge rep">R</span>'
    return '<span class="badge neutral">I</span>'


def build_table_rows(data, rank_key="rank_long"):
    rows = []
    for d in data:
        rank = d.get(rank_key, d.get("rank_short", ""))
        name = d.get("fullName", "")
        party = party_badge(d.get("party", ""))
        chamber = d.get("chamber", "")
        state = d.get("state", "")
        composite = fmt_float(d.get("composite"), 3)
        trades = d.get("trade_count", "")
        buys = d.get("buy_count", "")
        sells = d.get("sell_count", "")
        alpha = fmt_pct(d.get("mean_alpha_20d"))
        hit = fmt_pct(d.get("hit_rate"))
        sharpe = fmt_float(d.get("sharpe_alpha"), 2)
        lag = fmt_float(d.get("mean_lag_days"), 1)
        liq = fmt_pct(d.get("liq_pass_rate"))
        # Disclosure-lag cost column (ROADMAP #4). Signed: negative
        # means the follower's post-file entry captured less than the
        # member's trade-date anchor — expected sign for informed
        # trades. Backward-compatible with pre-#4 xlsx: missing → '—'.
        drag = fmt_pct(d.get("disclosure_drag_20d"))
        # Signal-quality filter columns (ROADMAP #3). Backward-compatible
        # with pre-#3 xlsx files: missing columns surface as None → '—'.
        non_self = fmt_pct(d.get("non_self_share"))
        late = fmt_pct(d.get("late_share"))
        etf_drops = int(d.get("etf_drops") or 0)
        opt_drops = int(d.get("options_drops") or 0)

        # Divergence flag
        div = d.get("divergence_flag")
        div_html = ' <span class="badge warn" title="Rank diverges between windows">⚠</span>' if div else ""

        rows.append(
            f"<tr>"
            f"<td class='rank-num'>{rank}</td>"
            f"<td>{name}{div_html}</td>"
            f"<td>{party}</td>"
            f"<td>{chamber}</td>"
            f"<td>{state}</td>"
            f"<td class='composite'>{composite}</td>"
            f"<td>{trades}</td>"
            f"<td class='green'>{buys}</td>"
            f"<td class='red'>{sells}</td>"
            f"<td>{alpha}</td>"
            f"<td>{drag}</td>"
            f"<td>{hit}</td>"
            f"<td>{sharpe}</td>"
            f"<td>{lag}d</td>"
            f"<td>{liq}</td>"
            f"<td>{non_self}</td>"
            f"<td>{late}</td>"
            f"<td>{etf_drops}</td>"
            f"<td>{opt_drops}</td>"
            f"</tr>"
        )
    return "\n          ".join(rows)


def build_weights_section(wb):
    if "Weights" not in wb.sheetnames:
        return ""
    _, data = read_sheet(wb, "Weights", max_rows=20)
    items = []
    for d in data:
        factor = d.get("factor", "").replace("_z", "").replace("_", " ").title()
        weight = d.get("weight", 0)
        pct = f"{weight * 100:.0f}%"
        items.append(f"<li><strong>{factor}</strong>: {pct}</li>")
    return "<ul>" + "".join(items) + "</ul>"


def build_benchmark_block(long_returns, short_returns):
    """Render a compact benchmark-returns card (4 tickers × 2 windows).

    Returns are gross of expense ratios and slippage; None → '—' via
    fmt_pct. The follow list is also reported gross, so the comparison
    is consistent (see design/benchmark-row.md, locked decision 5).
    """
    rows = []
    for ticker in BENCHMARK_TICKERS:
        long_r = fmt_pct(long_returns.get(ticker))
        short_r = fmt_pct(short_returns.get(ticker))
        rows.append(
            f"<tr><td><strong>{ticker}</strong></td>"
            f"<td>{long_r}</td><td>{short_r}</td></tr>"
        )
    return (
        '<div class="weights-card">'
        '<h3>Benchmark Reference — Cumulative Return</h3>'
        '<table style="font-size:0.82rem; width:auto;">'
        '<thead><tr><th>Ticker</th><th>365d</th><th>180d</th></tr></thead>'
        '<tbody>' + "".join(rows) + '</tbody>'
        '</table>'
        '</div>'
    )


PAGE_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Member Leaderboard — Congressional Trade Intelligence</title>
<style>
  :root {{
    --bg: #0f1117; --surface: #1a1d27; --surface2: #22263a;
    --accent: #4f8ef7; --green: #22c55e; --red: #ef4444;
    --yellow: #f59e0b; --text: #e2e8f0; --muted: #64748b;
    --border: #2e3450;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', system-ui, sans-serif;
    background: var(--bg); color: var(--text);
    padding: 32px 24px; max-width: 1400px; margin: 0 auto;
  }}
  a {{ color: var(--accent); text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  h1 {{ font-size: 1.6rem; font-weight: 700; color: var(--accent); margin-bottom: 4px; }}
  .breadcrumb {{ font-size: 0.82rem; color: var(--muted); margin-bottom: 24px; }}
  .subtitle {{ font-size: 0.88rem; color: var(--muted); margin-bottom: 28px; line-height: 1.6; }}
  .section-title {{
    font-size: 1.05rem; font-weight: 700; color: var(--accent);
    margin: 28px 0 14px; padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
  }}
  .table-wrap {{ width: 100%; overflow-x: auto; -webkit-overflow-scrolling: touch; margin-bottom: 28px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.82rem; }}
  th {{
    text-align: left; padding: 8px 10px;
    background: var(--surface2); color: var(--muted);
    font-weight: 600; font-size: 0.73rem;
    text-transform: uppercase; letter-spacing: 0.04em;
    position: sticky; top: 0;
  }}
  td {{ padding: 7px 10px; border-bottom: 1px solid var(--border); vertical-align: middle; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover {{ background: rgba(79,142,247,0.04); }}
  .rank-num {{ font-weight: 800; color: var(--muted); }}
  .composite {{ font-weight: 700; color: var(--accent); }}
  .green {{ color: var(--green); }}
  .red {{ color: var(--red); }}
  .badge {{ display: inline-block; font-size: 0.7rem; font-weight: 700; padding: 2px 7px; border-radius: 99px; }}
  .badge.dem {{ background: rgba(59,130,246,0.15); color: #60a5fa; }}
  .badge.rep {{ background: rgba(239,68,68,0.15); color: #f87171; }}
  .badge.neutral {{ background: rgba(100,116,139,0.15); color: var(--muted); }}
  .badge.warn {{ background: rgba(245,158,11,0.15); color: var(--yellow); font-size: 0.65rem; padding: 1px 5px; }}
  .weights-card {{
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 10px; padding: 18px; margin-bottom: 28px;
    font-size: 0.84rem;
  }}
  .weights-card h3 {{ font-size: 0.85rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 10px; }}
  .weights-card ul {{ list-style: none; display: flex; flex-wrap: wrap; gap: 12px 24px; }}
  .weights-card li {{ color: var(--text); }}
  .weights-card strong {{ color: var(--accent); }}
  .footer {{
    font-size: 0.75rem; color: var(--muted); margin-top: 32px;
    padding: 14px; background: var(--surface); border-radius: 8px; line-height: 1.6;
  }}
  @media (max-width: 720px) {{
    body {{ padding: 16px; }}
    h1 {{ font-size: 1.3rem; }}
    table {{ font-size: 0.75rem; white-space: nowrap; }}
    th, td {{ padding: 5px 6px; }}
  }}
</style>
</head>
<body>

<p class="breadcrumb"><a href="index.html">&larr; Back to Reports</a></p>
<h1>Member Followability Leaderboard</h1>
<p class="subtitle">
  Members of Congress ranked by trading followability — a composite score based on
  historical alpha, hit rate, signal timing, and liquidity. Updated monthly from
  public STOCK Act filings. Data as of {data_date}.
</p>

<div class="weights-card">
  <h3>Composite Scoring Weights</h3>
  {weights_html}
</div>

{benchmark_block}

<div class="section-title">365-Day Window (Top 20)</div>
<div class="table-wrap">
  <table>
    <thead><tr>
      <th>#</th><th>Member</th><th>Party</th><th>Chamber</th><th>State</th>
      <th>Score</th><th>Trades</th><th>Buys</th><th>Sells</th>
      <th title="Mean 20-business-day market-adjusted return on BUYs, entered at max(txDate, published) + 2 business days (ROADMAP #4). This is what a follower could capture — not what the member captured.">Mean &alpha; 20d</th>
      <th title="Disclosure drag: post-file mean alpha minus trade-date mean alpha. Negative means the follower missed runup to the publication lag — expected sign for informed trades.">Drag 20d</th>
      <th>Hit Rate</th><th>Sharpe</th><th>Lag</th><th>Liq Pass</th>
      <th title="Share of trades filed by a non-self owner (spouse, child, joint, dependent).">Non-self</th>
      <th title="Share of trades filed at day 40+ (STOCK Act deadline is 45 days).">Late</th>
      <th title="Count of broad-market ETF trades dropped from scoring.">ETF drops</th>
      <th title="Count of options / non-BUY-SELL trades dropped from scoring.">Opt drops</th>
    </tr></thead>
    <tbody>
      {long_rows}
    </tbody>
  </table>
</div>

<div class="section-title">180-Day Window (Top 20)</div>
<div class="table-wrap">
  <table>
    <thead><tr>
      <th>#</th><th>Member</th><th>Party</th><th>Chamber</th><th>State</th>
      <th>Score</th><th>Trades</th><th>Buys</th><th>Sells</th>
      <th title="Mean 20-business-day market-adjusted return on BUYs, entered at max(txDate, published) + 2 business days (ROADMAP #4). This is what a follower could capture — not what the member captured.">Mean &alpha; 20d</th>
      <th title="Disclosure drag: post-file mean alpha minus trade-date mean alpha. Negative means the follower missed runup to the publication lag — expected sign for informed trades.">Drag 20d</th>
      <th>Hit Rate</th><th>Sharpe</th><th>Lag</th><th>Liq Pass</th>
      <th title="Share of trades filed by a non-self owner (spouse, child, joint, dependent).">Non-self</th>
      <th title="Share of trades filed at day 40+ (STOCK Act deadline is 45 days).">Late</th>
      <th title="Count of broad-market ETF trades dropped from scoring.">ETF drops</th>
      <th title="Count of options / non-BUY-SELL trades dropped from scoring.">Opt drops</th>
    </tr></thead>
    <tbody>
      {short_rows}
    </tbody>
  </table>
</div>

<div class="footer">
  Composite scores are z-score-weighted aggregates of per-trade post-file alpha
  (BUYs measured against SPY at 5/20/60-day horizons, entered at max(txDate,
  published) + 2 business days per ROADMAP #4), hit rate, Sharpe ratio of alpha,
  reporting lag, liquidity pass rate, and trade count. The "Drag 20d" column
  shows post-file mean alpha minus trade-date mean alpha — the share of member
  alpha a follower loses to the STOCK Act disclosure lag. Benchmark returns
  above are gross of expense ratios and slippage. This is for informational
  purposes only and does not constitute investment advice.
  <br>Last built: {build_time}
</div>

</body>
</html>
"""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="site", help="Output directory")
    parser.add_argument("--top", type=int, default=20, help="Number of members to show per window")
    args = parser.parse_args()

    xlsx_path = find_latest_leaderboard()
    if not xlsx_path:
        print("No leaderboard xlsx found in scoring/output/. Skipping leaderboard page.")
        return

    # Extract date from filename
    basename = os.path.basename(xlsx_path)
    date_part = basename.replace("leaderboard_", "").replace(".xlsx", "")
    from datetime import datetime
    try:
        data_date = datetime.strptime(date_part, "%Y%m%d").strftime("%b %d, %Y")
    except ValueError:
        data_date = date_part

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # 365-day
    _, long_data = read_sheet(wb, "365d Leaderboard", max_rows=args.top)
    long_rows = build_table_rows(long_data, rank_key="rank_long")

    # 180-day
    _, short_data = read_sheet(wb, "180d Leaderboard", max_rows=args.top)
    short_rows = build_table_rows(short_data, rank_key="rank_short")

    weights_html = build_weights_section(wb)

    today = date.today()
    long_returns = all_benchmark_returns(today - timedelta(days=365), today)
    short_returns = all_benchmark_returns(today - timedelta(days=180), today)
    benchmark_block = build_benchmark_block(long_returns, short_returns)

    out_dir = Path(PROJ / args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    html = PAGE_TEMPLATE.format(
        data_date=data_date,
        weights_html=weights_html,
        benchmark_block=benchmark_block,
        long_rows=long_rows,
        short_rows=short_rows,
        build_time=datetime.now().strftime("%b %d, %Y at %I:%M %p"),
    )

    out_path = out_dir / "leaderboard.html"
    out_path.write_text(html)
    print(f"Built {out_path} ({len(long_data)} long / {len(short_data)} short rows)")


if __name__ == "__main__":
    main()
